import requests
import argparse
import sys
import xlrd
import xlwt
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument('--login_auto')
args = parser.parse_args()

if args.login_auto is None:
    sys.exit("login_auto 参数为空")


def get_inventory_data(goods_no):
    inventory_data = {'dc': '', 'qty': 0, 'ctry': ''}

    url = 'https://www.netcomponents.com/results.htm?flts=1&t=f&sm=&r=1&lgc=begins&pn1=' + goods_no
    cookie = {"login_name": "KevinQin", "login_org": "1085867", "login_user": "2", 'login_auto': args.login_auto}
    try:
        rep = requests.get(url, cookies=cookie)
        soup = BeautifulSoup(rep.text, 'lxml')
    except Exception as e:
        assert Exception("登录异常，请重新设置登录参数")

    all_inventory = soup.select(".div_table_float_reg")

    for inventory_area in all_inventory:
        inventory = inventory_area.select(".div_table_float_brkrd")[0]

        inventory_type = inventory.select(".partsrch_results.std_list tr.starttxt th")[0].text
        if inventory_type != 'In-Stock Inventory':
            continue

        ctrys = inventory.select(".ctry")
        qtys = inventory.select(".qty")
        dcs = inventory.select(".dc")
        sups = inventory.select(".sup")
        for i, ctry in enumerate(ctrys):
            if ctry.string == 'CN' or ctry.string == 'HK':
                continue
            authorized = sups[i].select('span[title="Authorized"]')
            if authorized:
                continue
            sup_text = sups[i].select(".suplink.lnk")[0].text
            if sup_text in ['FABtronics Pte Ltd', 'Chip 1 Exchange', 'SLCC Tech Inc', 'America Ⅱ Electronics', 'Newark Electronics', 'Farnell (F)']:
                continue
            if int(qtys[i].string) > inventory_data['qty']:
                inventory_data['dc'] = dcs[i].string
                inventory_data['qty'] = int(qtys[i].string)
                inventory_data['ctry'] = ctry.string

    return inventory_data


def main1():
    while True:
        try:
            filename = "D:/netcomponents/NXP.xlsx"
            wb = load_workbook(filename)  # 获取已存在的工作簿
            ws = wb.active  # 获取工作表
            nrows = ws.max_row
            for i in range(2, nrows+1):
                print(i, ws.cell(i, 2).value)
                if ws.cell(i, 13).value:
                    print(i, "已执行，跳过")
                    if i == nrows:
                        print("执行完毕")
                        sys.exit()
                    continue
                inventory_data = get_inventory_data(ws.cell(i, 2).value)
                ws.cell(i, 12, inventory_data['ctry'])
                ws.cell(i, 13, inventory_data['qty'])
                ws.cell(i, 14, inventory_data['dc'])
                wb.save(filename)
                if i == nrows:
                    print("执行完毕")
                    sys.exit()
        except Exception as e:
            print("执行异常", e)


def main():
    workbook = xlwt.Workbook(encoding='ascii')
    worksheet = workbook.add_sheet("Sheet1")

    try:
        filename = "D:/netcomponents/NXP.xls"
        data = xlrd.open_workbook(filename)
        table = data.sheets()[0]
        nrows = table.nrows
        for i in range(1, nrows):
            row_data = table.row(i)
            print(i, row_data[1].value)
            inventory_data = get_inventory_data(row_data[1].value)

            worksheet.write(i, 0, row_data[0].value)
            worksheet.write(i, 1, row_data[1].value)
            worksheet.write(i, 2, row_data[2].value)
            worksheet.write(i, 3, row_data[3].value)
            worksheet.write(i, 4, row_data[4].value)
            worksheet.write(i, 5, row_data[5].value)
            worksheet.write(i, 6, row_data[6].value)
            worksheet.write(i, 7, row_data[7].value)
            worksheet.write(i, 8, row_data[8].value)
            worksheet.write(i, 9, row_data[9].value)
            worksheet.write(i, 10, row_data[10].value)
            worksheet.write(i, 11, inventory_data['ctry'])
            worksheet.write(i, 12, inventory_data['qty'])
            worksheet.write(i, 13, inventory_data['dc'])
    except Exception as e:
        print("执行异常", e)

    filename = "D:/netcomponents/NXP_NEW" + time.strftime('%Y%m%d%H%M%S') + ".xls"
    workbook.save(filename)


if __name__ == '__main__':
    # main()
    main1()
